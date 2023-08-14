import pandas as pd
import numpy as np
import openpyxl as op

alpha = np.deg2rad(36.795)
gamma = np.deg2rad(90 - 78.169)
# 计算旋转矩阵
Rz = np.array([[np.cos(alpha),  np.sin(alpha), 0],
               [-np.sin(alpha), np.cos(alpha), 0],
               [0,              0,             1]])

Ry = np.array([[np.cos(gamma), 0, -np.sin(gamma)],
               [0,             1,              0],
               [np.sin(gamma), 0,  np.cos(gamma)]])

R = Ry @ Rz
R_inv = np.linalg.inv(R)

data = pd.read_csv('data1.csv', usecols=range(0, 4), encoding="gbk")

a_dict = {}

# i表示数据表中的行索引
i = 0

for row in data.iloc:
    x0 = row[1]
    y0 = row[2]
    z0 = row[3]
    # 求出每点在新坐标系下的坐标，并保存至a_dict
    A_0 = R @ [[x0], [y0], [z0]]
    a_dict[i] = A_0
    i = i + 1

# 新建一个存放b所对应的促动器变化值总和的字典
b_dict = {}

# b在[-301, -299.8]的范围之间变动，把0.001作为变化步长进行遍历：
for b in np.arange(-301, -299.8, 0.001):

    # 针对第一问的条件，由 f + CP = |b| 可得a关于b的一次函数关系
    a = 1 / (4 * (- b - 0.534 * 300.4))

    """
    在yOz平面中研究促动器伸缩量，取z = a * y ^ 2 + b平面进行研究
    通过观察，发现理想抛物面上的点到原点C的距离d满足从顶点A0'到z = -1/2a的点先变小，
    再到工作态抛物面边界的点(满足sqrt(x^2 + y^2) = 150)又变大的单调性关系
    因此确定筛选条件：1. sqrt(z^2 + y^2) - R >= -0.6   ,  z = -1/2a
                    2. sqrt(z^2 + y^2) - R <= +0.6   ,  y = 150
    """
    if (-b/a - 1/(4*a**2))**0.5 - 300.4 < -0.6 or (150**2 + (a*150**2+b)**2)**0.5 - 300.4 > 0.6:
        continue
    else:
        b_dict[b] = 0
        for j in range(0, i):
            x0 = a_dict[j][0][0]
            y0 = a_dict[j][1][0]
            z0 = a_dict[j][2][0]
            # 仅取工作态范围内的点  计算新的主索节点的纵坐标z1
            if x0**2+y0**2 <= 22591:
                """
                先转化为求平面上的交点的问题
                联立：1. z = a*y^2 + b
                     2. z = z0/sqrt(x0^2 + y0^2) * y
                解得关于z的一元二次方程 Az^2 + Bz + C = 0
                """
                A = a*((x0*x0+y0*y0)/(z0*z0))
                B = -1
                C = b
                if A != 0:
                    # 所求的点在xOy平面以下，满足z<0，故取较小的z1
                    z1 = (-B-(B*B-4*A*C)**0.5)/(2*A)
                    x1 = (z1/z0) * x0
                    y1 = (z1/z0) * y0
                    # 根据相似关系 d = R - R(z1/z0)
                    d = 300.4*(1-z1/z0)

                    if x1 ** 2 + y1 ** 2 <= 22500:
                        b_dict[b] += abs(d)
                else:
                    # 最终加上顶点的变化量(因为顶点的xy坐标均为0，需要单独计算)
                    b_dict[b] += abs(300.4 + b)

# final_b为最终确定的最优参数
final_b = 0
min_sum = np.inf
# 遍历字典求得最小变化量所对应的b
for (key, value) in b_dict.items():
    if value < min_sum:
        min_sum = value
        final_b = key
final_a = 1/(4*(-final_b - 0.534*300.4))

print(f"z = {round(final_a,6)}(x^2+y^2){round(final_b,3)}")


# 写.xlsx文件前的准备工作
workbook = op.Workbook()
worksheet1 = workbook.active
worksheet1.title = "调整后主索节点编号及坐标"
worksheet2 = workbook.create_sheet(title="促动器顶端伸缩量")
worksheet3 = workbook.create_sheet(title="理想抛物面顶点坐标")
dat1 = [["节点编号", "X坐标（米）", "Y坐标（米）", "Z坐标（米）"]]
dat2 = [["对应主索节点编号", "伸缩量（米）"]]
A_3 = R_inv@[[0], [0], [-300.89]]
dat3 = [["X坐标（米）", "Y坐标（米）", "Z坐标（米）"]]
dat3.append([A_3[0][0], A_3[1][0], A_3[2][0]])

'''
新坐标系下点(x0, y0, z0)对应的直线参数方程为x=x0*t, y=y0*t, z=z0*t
与求得的抛物线z = 0.00178(x^2+y^2)-300.89联立可得一元二次方程：
(0.00178*(x0**2 + y0**2)) * x^2+ (-z0) * x + (-300.89) = 0
即可解出交点坐标
'''
for j in range(0, i):
    x0 = a_dict[j][0][0]
    y0 = a_dict[j][1][0]
    z0 = a_dict[j][2][0]
    # 求解一元二次方程
    D = 0.00178*(x0**2 + y0**2)
    E = -z0
    F = -300.89
    if D != 0:
        t = (-E + (E**2 - 4 * D * F)**0.5)/(2*D)
        x1 = t*x0
        y1 = t*y0
        z1 = t*z0
        if x1**2+y1**2 <= 22500:
            A_1 = R_inv @ [[x1], [y1], [z1]]
            dat1.append([data.iloc[j][0], A_1[0][0], A_1[1][0], A_1[2][0]])
            dat2.append([data.iloc[j][0], (x1**2+y1**2+z1**2)**0.5-300.4])
    else:
        A_1 = R_inv @ [[0], [0], [final_b]]
        dat1.append([data.iloc[j][0], A_1[0][0], A_1[1][0], A_1[2][0]])
        dat2.append([data.iloc[j][0], (x1**2+y1**2+z1**2)**0.5-300.4])

# 将获得的结果写入result.xlsx文件
for row in dat1:
    worksheet1.append(row)
for row in dat2:
    worksheet2.append(row)
for row in dat3:
    worksheet3.append(row)
workbook.save(filename="result.xlsx")
